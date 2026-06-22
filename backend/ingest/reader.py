"""
Format-agnostic tabular reader for ITRM Jira exports.

Jira Service Management exports arrive as one of:
  - binary .xls  (OLE2 compound document)  -> xlrd
  - .xlsx        (zip/OOXML)                -> openpyxl
  - .csv / .tsv                             -> csv
  - .xls that is really an HTML <table>     -> BeautifulSoup

Returns: (headers: list[str], rows: list[list]) where every cell is a raw
Python value (str / float / None). Dates may come back as Excel serials
(float) or text — normalization happens later in normalize.py.
"""
from __future__ import annotations
import csv
import io


def detect_format(data: bytes, filename: str = "") -> str:
    name = (filename or "").lower()
    if data[:4] == b"\xD0\xCF\x11\xE0":
        return "xls"          # OLE2 binary
    if data[:2] == b"PK":
        return "xlsx"         # zip / OOXML
    head = data[:512].lstrip().lower()
    if head.startswith(b"<") or b"<table" in head or b"<html" in head:
        return "html"
    if name.endswith(".csv") or name.endswith(".tsv"):
        return "csv"
    # default: try csv
    return "csv"


def _read_xls(data: bytes):
    import xlrd
    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    rows = [sheet.row_values(r) for r in range(sheet.nrows)]
    if not rows:
        return [], []
    headers = [str(h).strip() for h in rows[0]]
    return headers, rows[1:]


def _read_xlsx(data: bytes):
    from openpyxl import load_workbook
    book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]
    rows = [[c for c in row] for row in sheet.iter_rows(values_only=True)]
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return headers, rows[1:]


def _read_csv(data: bytes):
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    delim = ","
    try:
        delim = csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
    except Exception:
        if sample.count(";") > sample.count(","):
            delim = ";"
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = list(reader)
    if not rows:
        return [], []
    headers = [str(h).strip() for h in rows[0]]
    return headers, rows[1:]


def _read_html(data: bytes):
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(data, "html.parser")
    table = soup.find("table")
    if not table:
        return [], []
    trs = table.find_all("tr")
    grid = []
    for tr in trs:
        cells = tr.find_all(["td", "th"])
        grid.append([c.get_text(strip=True) for c in cells])
    if not grid:
        return [], []
    headers = [str(h).strip() for h in grid[0]]
    return headers, grid[1:]


def read_table(data: bytes, filename: str = ""):
    fmt = detect_format(data, filename)
    if fmt == "xls":
        headers, rows = _read_xls(data)
    elif fmt == "xlsx":
        headers, rows = _read_xlsx(data)
    elif fmt == "html":
        headers, rows = _read_html(data)
    else:
        headers, rows = _read_csv(data)
    return fmt, headers, rows
